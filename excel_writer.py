"""
Excel Writer Module
Handles writing financial data to Excel template using label-based row lookup
Never relies on hardcoded row numbers
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import pandas as pd
from typing import Dict, List, Tuple, Optional
import io
from mapping import TEMPLATE_LABEL_MAPPING


def find_row_by_label(ws, label: str, search_column: int = 1, 
                      start_row: int = 1, end_row: int = 200) -> Optional[int]:
    """
    Find row number by searching for label in specified column
    
    Args:
        ws: Worksheet object
        label: Label text to search for
        search_column: Column to search (default 1 = Column A)
        start_row: Start searching from this row
        end_row: Stop searching at this row
    
    Returns:
        Row number if found, None otherwise
    """
    label_lower = label.lower().strip()
    
    for row in range(start_row, end_row + 1):
        cell_value = ws.cell(row, search_column).value
        if cell_value:
            cell_lower = str(cell_value).lower().strip()
            # Exact match or contains match
            if label_lower == cell_lower or label_lower in cell_lower:
                return row
    
    return None


def is_formula_cell(ws, row: int, col: int) -> bool:
    """
    Check if a cell contains a formula
    
    Args:
        ws: Worksheet object
        row: Row number
        col: Column number
    
    Returns:
        True if cell contains formula, False otherwise
    """
    cell = ws.cell(row, col)
    
    # Check if cell has a formula
    if hasattr(cell, 'data_type') and cell.data_type == 'f':
        return True
    
    # Also check the value
    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
        return True
    
    return False



def get_year_columns(ws,
                     data_years: List[int],
                     header_row: int = 2) -> Dict[int, int]:
    """
    Map data years to Excel columns by reading (and updating) the template's year header row.

    Template behavior (Year0-internal mode):
      - Statement years in columns B, C, D
      - Internal opening-balance Year 0 in column E (hidden/internal)

    Rules:
      - Statement years = latest 3 years present in the data
      - Year0 = (first statement year) - 1
      - Header row is updated accordingly
      - Returns {year -> column_number} based on the header row values
    """
    years_sorted = sorted({int(y) for y in data_years if pd.notna(y)})

    # Statement years = latest 3 available years (demo default)
    stmt_years = years_sorted[-3:] if len(years_sorted) >= 3 else years_sorted
    if not stmt_years:
        return {}

    year0 = int(stmt_years[0]) - 1

    stmt_cols = [2, 3, 4]  # B, C, D
    year0_col = 5          # E (internal)

    # Write statement year headers
    for i, y in enumerate(stmt_years):
        if i < len(stmt_cols):
            ws.cell(header_row, stmt_cols[i]).value = int(y)

    # Write internal Year0 header if column exists
    if ws.max_column >= year0_col:
        ws.cell(header_row, year0_col).value = int(year0)

    # Build mapping from header values
    year_col_map: Dict[int, int] = {}
    for col in range(2, min(ws.max_column, 20) + 1):
        v = ws.cell(header_row, col).value
        if isinstance(v, (int, float)) and 1900 <= int(v) <= 2200:
            year_col_map[int(v)] = col

    return year_col_map


def write_financial_data_to_template(template_path: str,
                                     financial_data: Dict[int, Dict],
                                     unit_scale: float = 1.0) -> io.BytesIO:
    """
    Write financial data to Excel template using label-based lookup
    
    Args:
        template_path: Path to template file
        financial_data: Dict of {year: {line_item: value}}
        unit_scale: Scale factor for amounts (1.0 for thousands, 1000.0 for dollars->thousands)
    
    Returns:
        BytesIO object containing the Excel file
    """
    # Load template
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # Get years and map to columns
    years = sorted(financial_data.keys())
    year_col_map = get_year_columns(ws, years, header_row=2)
    
    # Track what was written
    writes_performed = []
    warnings = []
    
    # Write data using label lookup
    for year, data in financial_data.items():
        if year not in year_col_map:
            warnings.append(f'Year {year} exceeds 3-year template limit, skipped')
            continue
        
        col = year_col_map[year]
        
        for data_key, value in data.items():
            # Find template label for this data key
            template_label = None
            for label, key in TEMPLATE_LABEL_MAPPING.items():
                if key == data_key:
                    template_label = label
                    break
            
            if not template_label:
                continue
            
            # Find row by label
            row = find_row_by_label(ws, template_label)
            
            if row is None:
                warnings.append(f'Label "{template_label}" not found in template')
                continue
            
            # Check if target cell is a formula
            if is_formula_cell(ws, row, col):
                warnings.append(f'Skipped {template_label} (formula cell)')
                continue
            
            # Apply scale and write value
            scaled_value = value / unit_scale if value else 0
            
            # Special handling for negatives (ensure correct signs)
            # Liabilities and equity should be positive
            # Expenses should be positive
            # Revenues should be positive
            
            ws.cell(row, col).value = scaled_value
            writes_performed.append(f'{template_label} ({year}): {scaled_value:.2f}')
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output



def calculate_financial_statements(df: pd.DataFrame,
                                   is_trial_balance: bool = True) -> Dict[int, Dict]:
    """
    Calculate financial statement line items from a *single* dataset.

    Notes:
    - Trial Balance (is_trial_balance=True): treated as a point-in-time snapshot dataset.
      We use the *latest TxnDate within each year* as the year-end balance sheet snapshot.
      Income statement items are NOT expected in TB-only mode (they will be 0 unless present).
    - GL Activity (is_trial_balance=False): treated as activity; we sum within each year.

    This function is kept for backward compatibility. For best results when you have
    BOTH TB (BS) and GL (IS), use calculate_3statements_from_tb_gl().
    """
    df = df.copy()
    df['TxnDate'] = pd.to_datetime(df['TxnDate'], errors='coerce')
    df = df[df['TxnDate'].notna()].copy()
    df['Year'] = df['TxnDate'].dt.year

    financial_data: Dict[int, Dict] = {}

    years = sorted(df['Year'].dropna().unique())

    for year in years:
        year_data = df[df['Year'] == year].copy()

        # TB: use the latest date in the year as the snapshot (avoids summing monthly snapshots)
        if is_trial_balance:
            latest_dt = year_data['TxnDate'].max()
            year_data = year_data[year_data['TxnDate'] == latest_dt].copy()

        def sum_category(category: str, debit_credit: str = 'both') -> float:
            cat_data = year_data[year_data['FSLI_Category'] == category]
            if cat_data.empty:
                return 0.0
            if debit_credit == 'debit':
                return float(cat_data['Debit'].sum())
            elif debit_credit == 'credit':
                return float(cat_data['Credit'].sum())
            else:
                # Net method (debit - credit)
                return float(cat_data['Debit'].sum() - cat_data['Credit'].sum())

        # Income Statement (GL expected; TB will be 0)
        revenue = sum_category('revenue', 'credit')
        cogs = sum_category('cogs', 'debit')
        distribution_expenses = sum_category('distribution_expenses', 'debit')
        marketing_admin = sum_category('marketing_admin', 'debit')
        research_dev = sum_category('research_dev', 'debit')
        depreciation_expense = sum_category('depreciation_expense', 'debit')
        interest_expense = sum_category('interest_expense', 'debit')
        tax_expense = sum_category('tax_expense', 'debit')  # optional; template can compute taxes

        gross_profit = revenue - cogs
        total_opex = distribution_expenses + marketing_admin + research_dev + depreciation_expense
        ebit = gross_profit - total_opex
        ebt = ebit - interest_expense
        net_income = ebt - tax_expense

        # Balance Sheet (TB expected; GL will usually be 0)
        cash = sum_category('cash')
        accounts_receivable = sum_category('accounts_receivable')
        inventory = sum_category('inventory')
        prepaid_expenses = sum_category('prepaid_expenses')
        other_current_assets = sum_category('other_current_assets')

        ppe_gross = sum_category('ppe_gross')
        accumulated_depreciation = abs(sum_category('accumulated_depreciation'))  # positive number
        # Liabilities (ensure positive)
        accounts_payable = abs(sum_category('accounts_payable'))
        accrued_payroll = abs(sum_category('accrued_payroll'))
        deferred_revenue = abs(sum_category('deferred_revenue'))
        interest_payable = abs(sum_category('interest_payable'))
        other_current_liabilities = abs(sum_category('other_current_liabilities'))
        income_taxes_payable = abs(sum_category('income_taxes_payable'))
        long_term_debt = abs(sum_category('long_term_debt'))

        # Equity
        common_stock = abs(sum_category('common_stock'))
        retained_earnings = abs(sum_category('retained_earnings'))

        financial_data[year] = {
            # Income Statement
            'revenue': revenue,
            'cogs': cogs,
            'distribution_expenses': distribution_expenses,
            'marketing_admin': marketing_admin,
            'research_dev': research_dev,
            'depreciation_expense': depreciation_expense,
            'interest_expense': interest_expense,
            'tax_expense': tax_expense,
            'net_income': net_income,

            # Balance Sheet
            'cash': cash,
            'accounts_receivable': accounts_receivable,
            'inventory': inventory,
            'prepaid_expenses': prepaid_expenses,
            'other_current_assets': other_current_assets,
            'ppe_gross': ppe_gross,
            'accumulated_depreciation': accumulated_depreciation,

            'accounts_payable': accounts_payable,
            'accrued_payroll': accrued_payroll,
            'deferred_revenue': deferred_revenue,
            'interest_payable': interest_payable,
            'other_current_liabilities': other_current_liabilities,
            'income_taxes_payable': income_taxes_payable,
            'long_term_debt': long_term_debt,

            'common_stock': common_stock,
            'retained_earnings': retained_earnings,
        }

    # Add simple cash flow items if multiple years are present (best-effort)
    if len(years) >= 2:
        for i in range(1, len(years)):
            y = years[i]
            p = years[i-1]
            cur = financial_data[y]
            pri = financial_data[p]

            delta_ar = cur['accounts_receivable'] - pri['accounts_receivable']
            delta_inventory = cur['inventory'] - pri['inventory']
            delta_prepaid = cur['prepaid_expenses'] - pri['prepaid_expenses']
            delta_other_current_assets = cur['other_current_assets'] - pri['other_current_assets']
            delta_ap = cur['accounts_payable'] - pri['accounts_payable']
            delta_accrued_payroll = cur['accrued_payroll'] - pri['accrued_payroll']
            delta_deferred_revenue = cur['deferred_revenue'] - pri['deferred_revenue']
            delta_interest_payable = cur['interest_payable'] - pri['interest_payable']
            delta_other_current_liabilities = cur['other_current_liabilities'] - pri['other_current_liabilities']
            delta_income_taxes_payable = cur['income_taxes_payable'] - pri['income_taxes_payable']
            delta_debt = cur['long_term_debt'] - pri['long_term_debt']

            capex = -(cur['ppe_gross'] - pri['ppe_gross'])
            stock_issuance = cur['common_stock'] - pri['common_stock']

            financial_data[y].update({
                'delta_ar': -delta_ar,
                'delta_inventory': -delta_inventory,
                'delta_prepaid': -delta_prepaid,
                'delta_other_current_assets': -delta_other_current_assets,
                'delta_ap': delta_ap,
                'delta_accrued_payroll': delta_accrued_payroll,
                'delta_deferred_revenue': delta_deferred_revenue,
                'delta_interest_payable': delta_interest_payable,
                'delta_other_current_liabilities': delta_other_current_liabilities,
                'delta_income_taxes_payable': delta_income_taxes_payable,
                'delta_debt': delta_debt,
                'capex': capex,
                'stock_issuance': stock_issuance,
            })

    return financial_data



def calculate_3statements_from_tb_gl(tb_df: pd.DataFrame,
                                    gl_df: pd.DataFrame,
                                    statement_years: int = 3) -> Dict[int, Dict]:
    """
    3-statement calculation using BOTH datasets with Year 0 internal-only opening balances.

    Rules (Strict-mode friendly):
      - TB is a BS snapshot dataset (we use the latest TxnDate within each year as year-end).
      - GL is an activity dataset (we sum within each year).
      - Output statement years = latest `statement_years` years present in TB/GL (typically 3).
      - Requires an opening snapshot Year0 = (first statement year - 1) present in TB.
      - Year0 is used ONLY internally (opening balances); users see only statement years.

    Returns:
        Dict[year, line_item->value] including Year0 + statement years.
        Year0 contains BS inputs only.
    """
    tb_fin = calculate_financial_statements(tb_df, is_trial_balance=True) if tb_df is not None else {}
    gl_fin = calculate_financial_statements(gl_df, is_trial_balance=False) if gl_df is not None else {}

    all_years = sorted(set(tb_fin.keys()) | set(gl_fin.keys()))
    if not all_years:
        return {}

    stmt_years = all_years[-int(statement_years):] if len(all_years) >= statement_years else all_years
    stmt_years = sorted(stmt_years)

    # Year0 requirement
    first_stmt_year = int(stmt_years[0])
    year0 = first_stmt_year - 1
    if year0 not in tb_fin:
        raise ValueError(
            f"Missing Year 0 opening snapshot in TB. "
            f"Expected prior-year year-end ({year0}-12-31 or latest date in {year0}) "
            f"to support opening balances for first statement year {first_stmt_year}."
        )

    years_out = [year0] + stmt_years

    # Keys we allow from GL (Income Statement only) to avoid overwriting TB snapshot BS values
    is_keys = [
        'revenue','cogs','distribution_expenses','marketing_admin','research_dev',
        'depreciation_expense','interest_expense','tax_expense','net_income'
    ]

    # Merge TB (BS) + GL (IS) into one dict per year
    combined: Dict[int, Dict] = {}
    for y in years_out:
        combined[y] = {}
        combined[y].update(tb_fin.get(y, {}))
        # Only statement years should pull GL activity
        if y in stmt_years:
            # Only bring Income Statement keys from GL to avoid overwriting BS snapshot values
            gl_part = gl_fin.get(y, {})
            combined[y].update({k: gl_part.get(k, 0.0) for k in is_keys})

    # Ensure required BS keys exist for all years (including Year0)
    bs_keys = [
        'cash','accounts_receivable','inventory','prepaid_expenses','other_current_assets',
        'ppe_gross','accumulated_depreciation',
        'accounts_payable','accrued_payroll','deferred_revenue','interest_payable',
        'other_current_liabilities','income_taxes_payable','long_term_debt',
        'common_stock','retained_earnings'
    ]
    for y in years_out:
        for k in bs_keys:
            combined[y].setdefault(k, 0.0)
        # For statement years, ensure IS keys exist (Year0 IS inputs stay empty/0)
        if y in stmt_years:
            for k in is_keys:
                combined[y].setdefault(k, 0.0)
        # For statement years, ensure IS keys exist (Year0 IS inputs stay empty/0)
        if y in stmt_years:
            for k in is_keys:
                combined[y].setdefault(k, 0.0)

    # Dividends via retained earnings roll-forward (statement years only)
    for i, y in enumerate(stmt_years):
        prev = year0 if i == 0 else stmt_years[i-1]
        re_begin = combined[prev].get('retained_earnings', 0.0)
        re_end = combined[y].get('retained_earnings', 0.0)
        ni = combined[y].get('net_income', 0.0)
        div = max(0.0, re_begin + ni - re_end)
        combined[y]['dividends'] = div

    # Cash-flow drivers from BS deltas (statement years only; Year0 is opening)
    wc_keys = [
        ('accounts_receivable','delta_ar', -1.0),
        ('inventory','delta_inventory', -1.0),
        ('prepaid_expenses','delta_prepaid', -1.0),
        ('other_current_assets','delta_other_current_assets', -1.0),
        ('accounts_payable','delta_ap', +1.0),
        ('accrued_payroll','delta_accrued_payroll', +1.0),
        ('deferred_revenue','delta_deferred_revenue', +1.0),
        ('interest_payable','delta_interest_payable', +1.0),
        ('other_current_liabilities','delta_other_current_liabilities', +1.0),
        ('income_taxes_payable','delta_income_taxes_payable', +1.0),
    ]

    for i, y in enumerate(stmt_years):
        prev = year0 if i == 0 else stmt_years[i-1]
        cur = combined[y]
        pri = combined[prev]

        for src, dst, sign in wc_keys:
            cur[dst] = sign * (cur.get(src, 0.0) - pri.get(src, 0.0))

        # Financing deltas
        cur['delta_debt'] = cur.get('long_term_debt', 0.0) - pri.get('long_term_debt', 0.0)
        cur['stock_issuance'] = cur.get('common_stock', 0.0) - pri.get('common_stock', 0.0)

        # CapEx from PPE gross change (clean demo: no disposals)
        cur['capex'] = -(cur.get('ppe_gross', 0.0) - pri.get('ppe_gross', 0.0))

    return combined


def compute_reconciliation_checks(financial_data: Dict[int, Dict]) -> Dict[int, Dict[str, float]]:
    """
    Compute reconciliation checks in a way that matches the updated template logic.

    IMPORTANT DESIGN (per project constraints):
      - Balance Sheet check MUST use the TB equity balances (Common Stock + Retained Earnings from TB),
        because the template now treats equity balances as INPUTS (blue) rather than forcing a roll-forward.
      - We still compute a Retained Earnings roll-forward (RE_calc) for diagnostics only.
        Any difference is reported as 'retained_earnings_diff' (TB_RE - RE_calc).

    Returns a dict keyed by statement year with:
      - balance_sheet_check: Assets - (Liabilities + Equity_TB)
      - cashflow_check: Cash(TB) - EndingCash_calc
      - retained_earnings_calc: RE_prev + NI - Div (diagnostic)
      - retained_earnings_tb: RE from TB (input)
      - retained_earnings_diff: TB_RE - RE_calc (diagnostic)
    """
    if not financial_data:
        return {}

    years = sorted(financial_data.keys())
    if len(years) < 2:
        return {}

    year0 = years[0]
    stmt_years = years[1:]

    # Diagnostic RE roll-forward starting from TB Year0 RE
    re_calc = {year0: float(financial_data[year0].get('retained_earnings', 0.0) or 0.0)}
    checks: Dict[int, Dict[str, float]] = {}

    for i, y in enumerate(stmt_years):
        prev = year0 if i == 0 else stmt_years[i-1]
        ni = float(financial_data[y].get('net_income', 0.0) or 0.0)
        div = float(financial_data[y].get('dividends', 0.0) or 0.0)
        re_calc[y] = float(re_calc[prev] + ni - div)

        # Assets (modelled set)
        cash = float(financial_data[y].get('cash', 0.0) or 0.0)
        ar = float(financial_data[y].get('accounts_receivable', 0.0) or 0.0)
        inv = float(financial_data[y].get('inventory', 0.0) or 0.0)
        pre = float(financial_data[y].get('prepaid_expenses', 0.0) or 0.0)
        oca = float(financial_data[y].get('other_current_assets', 0.0) or 0.0)

        ppe_g = float(financial_data[y].get('ppe_gross', 0.0) or 0.0)
        acc_dep = float(financial_data[y].get('accumulated_depreciation', 0.0) or 0.0)
        net_ppe = ppe_g - acc_dep
        assets = cash + ar + inv + pre + oca + net_ppe

        # Liabilities
        ap = float(financial_data[y].get('accounts_payable', 0.0) or 0.0)
        accr = float(financial_data[y].get('accrued_payroll', 0.0) or 0.0)
        defrev = float(financial_data[y].get('deferred_revenue', 0.0) or 0.0)
        intpay = float(financial_data[y].get('interest_payable', 0.0) or 0.0)
        ocl = float(financial_data[y].get('other_current_liabilities', 0.0) or 0.0)
        taxpay = float(financial_data[y].get('income_taxes_payable', 0.0) or 0.0)
        debt = float(financial_data[y].get('long_term_debt', 0.0) or 0.0)
        liabilities = ap + accr + defrev + intpay + ocl + taxpay + debt

        # Equity (TB inputs, per template)
        cs = float(financial_data[y].get('common_stock', 0.0) or 0.0)
        re_tb = float(financial_data[y].get('retained_earnings', 0.0) or 0.0)
        equity_tb = cs + re_tb

        bs_check = assets - (liabilities + equity_tb)

        # Cashflow check (standard cash roll-forward)
        begin_cash = float(financial_data[prev].get('cash', 0.0) or 0.0)
        dep = float(financial_data[y].get('depreciation_expense', 0.0) or 0.0)

        wc = (
            float(financial_data[y].get('delta_ar', 0.0) or 0.0) +
            float(financial_data[y].get('delta_inventory', 0.0) or 0.0) +
            float(financial_data[y].get('delta_prepaid', 0.0) or 0.0) +
            float(financial_data[y].get('delta_other_current_assets', 0.0) or 0.0) +
            float(financial_data[y].get('delta_ap', 0.0) or 0.0) +
            float(financial_data[y].get('delta_accrued_payroll', 0.0) or 0.0) +
            float(financial_data[y].get('delta_deferred_revenue', 0.0) or 0.0) +
            float(financial_data[y].get('delta_interest_payable', 0.0) or 0.0) +
            float(financial_data[y].get('delta_other_current_liabilities', 0.0) or 0.0) +
            float(financial_data[y].get('delta_income_taxes_payable', 0.0) or 0.0)
        )
        capex = float(financial_data[y].get('capex', 0.0) or 0.0)
        stock = float(financial_data[y].get('stock_issuance', 0.0) or 0.0)
        deltadebt = float(financial_data[y].get('delta_debt', 0.0) or 0.0)

        net_cash_change = (ni + dep + wc) + capex + (stock - div + deltadebt)
        end_cash_calc = begin_cash + net_cash_change
        cf_check = cash - end_cash_calc

        checks[y] = {
            'balance_sheet_check': bs_check,
            'cashflow_check': cf_check,
            'retained_earnings_calc': re_calc[y],
            'retained_earnings_tb': re_tb,
            'retained_earnings_diff': re_tb - re_calc[y],
        }

    return checks



def validate_template_structure(template_path: str) -> Tuple[bool, List[str]]:
    """
    Validate that template has expected structure
    
    Returns:
        (is_valid, list_of_issues)
    """
    issues = []
    
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Check key labels exist
        required_labels = [
            'Revenues',
            'Cost of Goods Sold',
            'Cash',
            'Accounts Payable',
            'Common Stock and Additional Paid-In Capital',
        ]
        
        for label in required_labels:
            row = find_row_by_label(ws, label)
            if row is None:
                issues.append(f'Required label "{label}" not found')
        
        return len(issues) == 0, issues
        
    except Exception as e:
        issues.append(f'Error loading template: {str(e)}')
        return False, issues
