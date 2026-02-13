"""Three Statements Automation - Streamlit App (patched)
Focus:
- TB + GL must be treated as a set for generation.
- Validation section is always visible and detailed.
- Auto-fix is user-controlled (accept/reject) instead of silently applied.
- Demo controls:
  1) Download Sample Financial Model (SAMPLE_DEMO template)
  2) Download Sample Dataset (current loaded TB+GL as a zip)
  3) Load Random Sample Set (TB+GL together, matched years)
"""

import io
import zipfile
from datetime import datetime

import pandas as pd
import openpyxl
import numpy as np
import streamlit as st

from validation import (
    validate_trial_balance,
    validate_gl_activity,
    apply_auto_fixes,
    validate_year0_opening_snapshot,
    add_year0_snapshot,
)
from mapping import map_accounts, TEMPLATE_LABEL_MAPPING
from excel_writer import (
    calculate_3statements_from_tb_gl,
    calculate_financial_statements,
    write_financial_data_to_template,
    compute_reconciliation_checks,
)
from sample_data import get_sample_data_path, get_template_path
from pdf_export import generate_pdf_report
from ai_summary import generate_ai_summary
import os


# ----------------------------
# Config / defaults
# ----------------------------
st.set_page_config(page_title="AI Accounting Agent", layout="wide")

AUTO_FIX_OPTIONS = [
    ("remove_missing_dates", "Remove rows with missing TxnDate"),
    ("remove_future_dates", "Remove rows with future TxnDate"),
    ("map_unclassified", "Map missing AccountNumber → 9999"),
    ("fix_account_numbers", "Convert negative AccountNumber → positive"),
    ("remove_duplicates", "Remove fully duplicate rows (safe)"),
]

UNIT_SCALE_OPTIONS = {
    "USD (no scaling)": 1,
    "USD thousands (÷ 1000)": 1000,
}

# Session state init
for k, default in {
    "tb_df": None,
    "gl_df": None,
    "tb_name": None,
    "gl_name": None,
    "tb_changes": [],
    "gl_changes": [],
    "validation_tb": [],
    "validation_gl": [],
    "selected_fixes": [],
    "unit_scale": 1000,
    "strict_mode": True,
    "template_type": "zero",  # 'zero' or 'demo'
    "dataset_source": None,  # 'random' or 'upload'
    "preview_sections": {},
    "last_excel_bytes": None,
}.items():
    if k not in st.session_state:
        st.session_state[k] = default


def _issues_to_table(issues):
    """Compact table for issue list."""
    rows = []
    for it in issues:
        rows.append({
            "severity": it.get("severity", ""),
            "category": it.get("category", ""),
            "issue": it.get("issue", ""),
            "impact": it.get("impact", ""),
            "suggestion": it.get("suggestion", ""),
            "auto_fix": it.get("auto_fix", ""),
        })
    return pd.DataFrame(rows)


def _count_by_severity(issues):
    counts = {"Critical": 0, "Warning": 0, "Info": 0}
    for it in issues:
        sev = it.get("severity", "Info")
        if sev not in counts:
            counts[sev] = 0
        counts[sev] += 1
    return counts


def _find_row_exact(ws, label: str, start_row: int = 1, end_row: int = 200):
    target = str(label).strip().lower()
    for r in range(start_row, end_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        if str(v).strip().lower() == target:
            return r
    return None


def _extract_labels(ws, start_row: int, end_row: int) -> list[str]:
    labels = []
    for r in range(start_row, end_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            labels.append("")
        else:
            labels.append(str(v).strip())
    return labels


def _compute_template_preview_sections(financial_data: dict, template_path: str, unit_scale: int = 1):
    """
    Build template-matching preview tables (Income Statement / Balance Sheet / Cash Flow)
    using the template's label order as the source of truth.

    Returns:
        dict[str, pd.DataFrame] sections
    """
    if not financial_data:
        return {}

    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb["Blank 3 Statement Model"] if "Blank 3 Statement Model" in wb.sheetnames else wb[wb.sheetnames[0]]

    years_all = sorted(financial_data.keys())
    stmt_years = years_all[1:] if len(years_all) > 1 else years_all  # hide Year0 in preview

    # Helper: fetch value by internal key
    def v(y: int, key: str) -> float:
        try:
            return float(financial_data.get(y, {}).get(key, 0.0) or 0.0) / float(unit_scale)
        except Exception:
            return 0.0

    # Derived calculations (match template logic)
    def gross_profit(y): return v(y, "revenue") - v(y, "cogs")
    def total_opex(y): 
        return v(y,"distribution_expenses")+v(y,"marketing_admin")+v(y,"research_dev")+v(y,"depreciation_expense")
    def ebit(y): return gross_profit(y) - total_opex(y)
    def ebt(y): return ebit(y) - v(y,"interest_expense")
    def net_income(y): 
        # prefer calculated if present, else compute
        ni = v(y,"net_income")
        return ni if abs(ni) > 1e-9 else (ebt(y) - v(y,"tax_expense"))

    def total_current_assets(y):
        return v(y,"cash")+v(y,"accounts_receivable")+v(y,"inventory")+v(y,"prepaid_expenses")+v(y,"other_current_assets")
    def net_ppe(y):
        return v(y,"ppe_gross") - v(y,"accumulated_depreciation")
    def total_assets(y):
        return total_current_assets(y) + net_ppe(y)

    def total_current_liab(y):
        return (v(y,"accounts_payable")+v(y,"accrued_payroll")+v(y,"deferred_revenue")+
                v(y,"interest_payable")+v(y,"other_current_liabilities")+v(y,"income_taxes_payable"))
    def total_equity(y):
        return v(y,"common_stock") + v(y,"retained_earnings")
    def total_le(y):
        return total_current_liab(y) + v(y,"long_term_debt") + total_equity(y)

    # Cash Flow
    def cfo(y):
        return (net_income(y) + v(y,"depreciation_expense") +
                v(y,"delta_ar")+v(y,"delta_inventory")+v(y,"delta_prepaid")+v(y,"delta_other_current_assets")+
                v(y,"delta_ap")+v(y,"delta_accrued_payroll")+v(y,"delta_deferred_revenue")+v(y,"delta_interest_payable")+
                v(y,"delta_other_current_liabilities")+v(y,"delta_income_taxes_payable"))
    def cfi(y): return v(y,"capex")
    def cff(y): return v(y,"stock_issuance") + (-v(y,"dividends")) + v(y,"delta_debt")
    def net_change_cash(y): return cfo(y) + cfi(y) + cff(y)
    def begin_cash(y):
        bc = v(y, "beginning_cash")
        if bc != 0:
            return bc
        # Derive beginning cash from prior-year cash (Year0 for first statement year)
        if y in stmt_years:
            idx = stmt_years.index(y)
            prev = year0 if idx == 0 else stmt_years[idx - 1]
            return v(prev, "cash")
        return 0.0
    def end_cash(y):
        # prefer key if exists, else compute
        ec = v(y,"cash")
        return ec if abs(ec) > 1e-9 else (begin_cash(y) + net_change_cash(y))

    # Checks (match your Row3/Row81 intent)
    checks = compute_reconciliation_checks(financial_data)
    def bs_check(y): return float(checks.get(y, {}).get("balance_sheet_check", 0.0) or 0.0) / float(unit_scale)
    def cf_check(y): return float(checks.get(y, {}).get("cashflow_check", 0.0) or 0.0) / float(unit_scale)

    # Map template label -> internal key (reverse mapping)
    label_to_key = TEMPLATE_LABEL_MAPPING

    # Build a section DataFrame from template row range
    def build_df(row_start: int, row_end: int) -> pd.DataFrame:
        labels = _extract_labels(ws, row_start, row_end)
        data = {}
        for y in stmt_years:
            col = []
            for lab in labels:
                if lab == "":
                    col.append(np.nan)
                    continue


                # Direct mapped inputs
                if lab in label_to_key:
                    col.append(v(y, label_to_key[lab]))
                    continue

                # Derived / totals
                derived = {
                    "Gross Profit": gross_profit,
                    "EBIT (Operating Profit)": ebit,
                    "Income Before Taxes": ebt,
                    "Net Income": net_income,
                    "Total Current Assets": total_current_assets,
                    "Property Plant and Equipment - Net": net_ppe,
                    "TOTAL ASSETS": total_assets,
                    "Total Current Liabilities:": total_current_liab,
                    "Total Shareholders' Equity": total_equity,
                    "TOTAL LIABILITIES AND SHAREHOLDERS' EQUITY": total_le,
                    "Net Cash Provided by Operating Activities": cfo,
                    "Cash Flows from Investing Activities": cfi,
                    "Cash Flows from Financing Activities": cff,
                    "Increase/(Decrease) in Cash and Equivalents": net_change_cash,
                    "Cash and Equivalents, Beginning of the Year": begin_cash,
                    "Cash and Equivalents, End of the Year": end_cash,
                    # Checks (these rows are outside the sections, but safe)
                    "Balance Sheet Check (A - L + E)": bs_check,
                    "Check": cf_check,
                }
                if lab in derived:
                    col.append(float(derived[lab](y)))
                    continue


                # Headings / section labels should stay blank; unmapped numeric rows show 0 to avoid "missing statement" look
                heading_labels = {
                    "ASSETS", "LIABILITIES AND SHAREHOLDERS' EQUITY",
                    "Current Assets:", "Non-Current Assets:", "Current Liabilities:",
                    "Non-Current Liabilities:", "Shareholder's Equity:",
                    "Cash Flow Statement", "Cash Flows from Operating Activities:",
                    "Changes in Operating Assets and Liabilities:", "Investing Activities:",
                    "Financing Activities:",
                    "Revenues", "Operating Expenses", "Other Expense / (Income)", "Taxes",
                }
                is_heading = (lab in heading_labels) or (lab.endswith(":") and not lab.lower().startswith("total"))
                # Some templates use ALL CAPS for section headers; do NOT treat totals as headings (handled above in derived).
                is_heading = is_heading or (lab.upper() == lab and lab not in {"TOTAL ASSETS", "TOTAL LIABILITIES AND SHAREHOLDERS' EQUITY"})

                if is_heading:
                    col.append(np.nan)
                else:
                    col.append(0.0)

            data[f"FY{y}"] = col

        df = pd.DataFrame(data, index=labels)
        return df

    # Find row ranges by labels (use the template's second 'Income Statement' section for actual output)
    is_header = _find_row_exact(ws, "Income Statement", start_row=25, end_row=120)  # should find row 30
    is_start = _find_row_exact(ws, "Revenues", start_row=is_header or 1, end_row=140)
    is_end = _find_row_exact(ws, "Common Dividends", start_row=is_start or 1, end_row=160)

    bs_header = _find_row_exact(ws, "Balance Sheet", start_row=40, end_row=120)  # should find row 48
    bs_start = _find_row_exact(ws, "ASSETS", start_row=bs_header or 1, end_row=200)
    bs_end = _find_row_exact(ws, "Check", start_row=bs_start or 1, end_row=140)  # row 81 (check line)

    cf_header = _find_row_exact(ws, "Cash Flow Statement", start_row=70, end_row=160)  # row 84
    cf_start = _find_row_exact(ws, "Cash Flows from Operating Activities:", start_row=cf_header or 1, end_row=200)
    cf_end = _find_row_exact(ws, "Cash and Equivalents, End of the Year", start_row=cf_start or 1, end_row=220)

    sections = {}

    # Income Statement (include from Revenues down to Common Dividends)
    if is_start and is_end:
        sections["Income Statement"] = build_df(is_start, is_end)
    # Balance Sheet (include from ASSETS down to TOTAL L+E)
    if bs_start and bs_end:
        # bs_end is 'Check' row; include up to TOTAL L+E row (one above check)
        sections["Balance Sheet"] = build_df(bs_start, bs_end - 2)
    # Cash Flow (include from Operating Activities down to End of Year)
    if cf_start and cf_end:
        sections["Cash Flow Statement"] = build_df(cf_start, cf_end)

    # Checks section (explicit)
    checks_df = pd.DataFrame(
        {
            f"FY{y}": {
                "Balance Sheet Check (A - L + E)": bs_check(y),
                "Cash Tie-out Check": cf_check(y),
            }
            for y in stmt_years
        }
    )
    sections["Checks"] = checks_df

    return sections


def run_validation():
    """Run validations and store results in session state."""
    tb_df = st.session_state["tb_df"]
    gl_df = st.session_state["gl_df"]

    tb_issues = validate_trial_balance(tb_df) if tb_df is not None else []
    gl_issues = validate_gl_activity(gl_df) if gl_df is not None else []

    st.session_state["validation_tb"] = tb_issues
    st.session_state["validation_gl"] = gl_issues

    return tb_issues, gl_issues


def load_random_set():
    """Load matched TB + GL backup sets by choosing one year-range and loading both files."""
    year_ranges = [(2020, 2022), (2021, 2023), (2022, 2024), (2023, 2025), (2024, 2026)]
    import random
    y0, y1 = random.choice(year_ranges)
    tb_file = f"backup_tb_{y0}_{y1}.csv"
    gl_file = f"backup_gl_{y0}_{y1}_with_txnid.csv"

    tb_path = get_sample_data_path(tb_file)
    gl_path = get_sample_data_path(gl_file)

    tb_df = pd.read_csv(tb_path)
    gl_df = pd.read_csv(gl_path)

    st.session_state["tb_df"] = tb_df
    st.session_state["gl_df"] = gl_df
    st.session_state["tb_name"] = tb_file
    st.session_state["gl_name"] = gl_file
    st.session_state["tb_changes"] = []
    st.session_state["gl_changes"] = []
    st.session_state["dataset_source"] = "random"

    run_validation()


def dataset_zip_bytes(tb_df: pd.DataFrame, gl_df: pd.DataFrame, tb_name: str, gl_name: str) -> bytes:
    bio = io.BytesIO()
    with zipfile.ZipFile(bio, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(tb_name or "tb.csv", tb_df.to_csv(index=False))
        zf.writestr(gl_name or "gl.csv", gl_df.to_csv(index=False))
    bio.seek(0)
    return bio.read()


# ----------------------------
# Sidebar
# ----------------------------
with st.sidebar:
    st.header("Demo")

    # A1) Download sample financial model (demo template)
    demo_template_path = get_template_path("demo")
    try:
        with open(demo_template_path, "rb") as f:
            st.download_button(
                "Download Sample Financial Model",
                data=f.read(),
                file_name="Financial_Model_SAMPLE_DEMO_USD_thousands_GAAP.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception as e:
        st.caption(f"Could not load demo template: {e}")

    # A3) Load random sample set (TB+GL together)
    if st.button("Load Random Sample Set (TB + GL, 3 years)"):
        load_random_set()

    # A2) Download current dataset (TB+GL zip)
    if st.session_state["tb_df"] is not None and st.session_state["gl_df"] is not None:
        zbytes = dataset_zip_bytes(
            st.session_state["tb_df"],
            st.session_state["gl_df"],
            st.session_state["tb_name"] or "tb.csv",
            st.session_state["gl_name"] or "gl.csv",
        )
        st.download_button(
            "Download Sample Dataset (TB + GL)",
            data=zbytes,
            file_name="sample_dataset_tb_gl.zip",
            mime="application/zip",
        )
    else:
        st.caption("Load a dataset first to enable dataset download.")

    st.divider()
    st.header("Inputs")

    tb_up = st.file_uploader("Upload TB (CSV)", type=["csv"], key="tb_uploader")
    gl_up = st.file_uploader("Upload GL (CSV)", type=["csv"], key="gl_uploader")

    if tb_up is not None:
        st.session_state["tb_df"] = pd.read_csv(tb_up)
        st.session_state["tb_name"] = getattr(tb_up, "name", "tb.csv")
        st.session_state["tb_changes"] = []
        st.session_state["dataset_source"] = "upload"
        run_validation()

    if gl_up is not None:
        st.session_state["gl_df"] = pd.read_csv(gl_up)
        st.session_state["gl_name"] = getattr(gl_up, "name", "gl.csv")
        st.session_state["gl_changes"] = []
        st.session_state["dataset_source"] = "upload"
        run_validation()

    st.divider()
    st.header("Settings")

    # Unit scale (you asked for 2-level, not free numeric)
    unit_label = st.radio("Unit scale (divide by)", list(UNIT_SCALE_OPTIONS.keys()), index=1)
    st.session_state["unit_scale"] = UNIT_SCALE_OPTIONS[unit_label]

    st.session_state["strict_mode"] = st.checkbox("Strict mode", value=st.session_state["strict_mode"])

    # Template type kept but tucked away (you said the picker isn't very useful)
    with st.expander("Advanced: template selection"):
        template_label = st.radio("Excel template", ["ZERO (processing)", "SAMPLE (demo)"], index=0)
        st.session_state["template_type"] = "zero" if template_label.startswith("ZERO") else "demo"


# ----------------------------
# Main UI
# ----------------------------
st.title("AI Accounting Agent — GL → 3-Statement Demo")

tb_df = st.session_state["tb_df"]
gl_df = st.session_state["gl_df"]

col1, col2 = st.columns(2)
with col1:
    st.subheader("Trial Balance (TB)")
    if tb_df is None:
        st.info("No TB loaded yet.")
    else:
        st.caption(f"Loaded: {st.session_state['tb_name']}")
        st.dataframe(tb_df.head(20), use_container_width=True)

with col2:
    st.subheader("General Ledger (GL)")
    if gl_df is None:
        st.info("No GL loaded yet.")
    else:
        st.caption(f"Loaded: {st.session_state['gl_name']}")
        st.dataframe(gl_df.head(20), use_container_width=True)


# ----------------------------
# Validation + auto-fix (this is the section you said must NOT disappear)
# ----------------------------
tb_issues = st.session_state.get("validation_tb", []) or []
gl_issues = st.session_state.get("validation_gl", []) or []

tb_counts = _count_by_severity(tb_issues)
gl_counts = _count_by_severity(gl_issues)

banner_parts = []
if tb_df is not None:
    banner_parts.append(f"TB [Critical]: {tb_counts.get('Critical',0)} [Warning]: {tb_counts.get('Warning',0)} [Info]: {tb_counts.get('Info',0)}")
if gl_df is not None:
    banner_parts.append(f"GL [Critical]: {gl_counts.get('Critical',0)} [Warning]: {gl_counts.get('Warning',0)} [Info]: {gl_counts.get('Info',0)}")

if banner_parts:
    st.warning("Validation issues found: " + " | ".join(banner_parts))

with st.expander("Validation details", expanded=True):
    # TB
    st.markdown("### Trial Balance (TB)")
    if tb_df is None:
        st.caption("No TB loaded.")
    elif not tb_issues:
        st.success("No TB validation issues.")
    else:
        st.dataframe(_issues_to_table(tb_issues), use_container_width=True)
        for idx, it in enumerate(tb_issues[:20], start=1):
            with st.expander(f"TB Issue {idx}: {it.get('issue','')}"):
                st.markdown(f"**Severity:** {it.get('severity','')}  |  **Category:** {it.get('category','')}")
                if it.get("impact"): st.markdown(f"**Impact:** {it.get('impact')}")
                if it.get("suggestion"): st.markdown(f"**Suggestion:** {it.get('suggestion')}")
                if it.get("auto_fix"): st.markdown(f"**Auto-fix option:** `{it.get('auto_fix')}`")
                if "sample_data" in it and it["sample_data"] is not None:
                    st.dataframe(it["sample_data"], use_container_width=True)

    st.divider()

    # GL
    st.markdown("### General Ledger (GL)")
    if gl_df is None:
        st.caption("No GL loaded.")
    elif not gl_issues:
        st.success("No GL validation issues.")
    else:
        st.dataframe(_issues_to_table(gl_issues), use_container_width=True)
        for idx, it in enumerate(gl_issues[:20], start=1):
            with st.expander(f"GL Issue {idx}: {it.get('issue','')}"):
                st.markdown(f"**Severity:** {it.get('severity','')}  |  **Category:** {it.get('category','')}")
                if it.get("impact"): st.markdown(f"**Impact:** {it.get('impact')}")
                if it.get("suggestion"): st.markdown(f"**Suggestion:** {it.get('suggestion')}")
                if it.get("auto_fix"): st.markdown(f"**Auto-fix option:** `{it.get('auto_fix')}`")
                if "sample_data" in it and it["sample_data"] is not None:
                    st.dataframe(it["sample_data"], use_container_width=True)

    st.divider()

    # Auto-fix controls (accept/reject)
    st.markdown("### Auto-fix (accept/reject)")
    detected_fixes = sorted({it.get("auto_fix") for it in (tb_issues + gl_issues) if it.get("auto_fix")})
    if not detected_fixes:
        st.caption("No auto-fixes suggested by the current validation results.")
    else:
        fix_labels = {code: label for code, label in AUTO_FIX_OPTIONS}
        default_selected = [f for f in detected_fixes if f != "remove_duplicates"]  # safer default
        selected = st.multiselect(
            "Select fixes to apply",
            options=detected_fixes,
            default=default_selected,
            format_func=lambda x: f"{x} — {fix_labels.get(x, '')}",
        )

        cA, cB = st.columns([1, 1])
        with cA:
            if st.button("Apply selected fixes"):
                if tb_df is not None:
                    fixed, changes = apply_auto_fixes(tb_df, selected_fixes=selected)
                    st.session_state["tb_df"] = fixed
                    st.session_state["tb_changes"] = changes
                if gl_df is not None:
                    fixed, changes = apply_auto_fixes(gl_df, selected_fixes=selected)
                    st.session_state["gl_df"] = fixed
                    st.session_state["gl_changes"] = changes
                run_validation()
                st.success("Applied selected fixes and re-ran validation.")
        with cB:
            if st.button("Re-run validation"):
                run_validation()
                st.info("Re-ran validation.")

        if st.session_state.get("tb_changes"):
            st.caption("TB auto-fixes applied:")
            st.write(st.session_state["tb_changes"])
        if st.session_state.get("gl_changes"):
            st.caption("GL auto-fixes applied:")
            st.write(st.session_state["gl_changes"])


# ----------------------------
# Generation
# ----------------------------
st.divider()
st.header("Generate Outputs")

strict_mode = st.session_state["strict_mode"]

if st.button("Generate 3-Statement Outputs", type="primary"):
    # Require TB + GL as a set for generation
    if tb_df is None or gl_df is None:
        st.error("TB and GL must be loaded as a set to generate the 3-statement output.")
        st.stop()

    # Strict mode gate: fail if any Critical issues remain
    tb_issues, gl_issues = run_validation()
    crit = [it for it in (tb_issues + gl_issues) if it.get("severity") == "Critical"]
    if strict_mode and crit:
        st.error("Strict mode: Critical validation issues must be resolved before generation. See Validation details above.")
        st.stop()

    # Year0 strict requirement
    if strict_mode:
        year0_issues = validate_year0_opening_snapshot(tb_df, statement_years=3)
        if year0_issues:
            # If the dataset came from the built-in random demo/backups, we can create an internal Year0 snapshot
            # (you asked for Year0 to exist for internal use). For user uploads, we do NOT synthesize Year0.
            if st.session_state.get("dataset_source") == "random":
                tb_fixed, msg = add_year0_snapshot(tb_df, statement_years=3)
                st.session_state["tb_df"] = tb_fixed
                tb_df = tb_fixed
                st.info(msg)
                run_validation()
                year0_issues = validate_year0_opening_snapshot(tb_df, statement_years=3)

            if year0_issues:
                st.error("Strict mode: Year0 opening snapshot requirement failed:\n" + "\n".join(year0_issues))
                st.stop()

    # Map accounts
    tb_mapped = map_accounts(tb_df)
    gl_mapped = map_accounts(gl_df)

    # Compute 3-statement data
    financial_data = calculate_3statements_from_tb_gl(tb_mapped, gl_mapped)

    # Write to template
    template_path = get_template_path(st.session_state["template_type"])
    out_bytes = write_financial_data_to_template(
        template_path=template_path,
        financial_data=financial_data,
        unit_scale=float(st.session_state["unit_scale"]),
    )

    # Persist output so Streamlit reruns don’t lose it
    st.session_state["last_excel_bytes"] = out_bytes.getvalue()

    # Build a template-matching preview (Income Statement / Balance Sheet / Cash Flow)
    try:
        template_path_for_preview = get_template_path(st.session_state.get("template_type", "zero"))
        st.session_state["preview_sections"] = _compute_template_preview_sections(
            financial_data=financial_data,
            template_path=template_path_for_preview,
            unit_scale=st.session_state.get("unit_scale", 1),
        )
    except Exception:
        st.session_state["preview_sections"] = {}

    st.success("Generated Excel output.")
    st.download_button(
        "Download Excel Output",
        data=out_bytes.getvalue(),
        file_name="3statement_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ========================================
    # 3-Statement Preview
    # ========================================
    if st.session_state.get("preview_sections"):
        st.subheader("3-Statement Output Preview (Template-Matching)")

        sections = st.session_state["preview_sections"]
        tab_names = [k for k in ["Income Statement", "Balance Sheet", "Cash Flow Statement", "Checks"] if k in sections]
        if tab_names:
            tabs = st.tabs(tab_names)
            for tname, tab in zip(tab_names, tabs):
                with tab:
                    df = sections.get(tname)
                    if df is None or (hasattr(df, "empty") and df.empty):
                        st.info("No preview data available for this section.")
                    else:
                        # Clean display: Streamlit shows Python None as 'None' which is misleading for section headers.
                        # For website preview only: show blanks for headers and 0 for unmapped numeric rows (already 0 in df).
                        df_show = df.copy()
                        df_show = df_show.replace({None: np.nan})

                        def _fmt_cell(x):
                            if pd.isna(x):
                                return ""
                            if isinstance(x, (int, float, np.integer, np.floating)):
                                return f"{float(x):,.0f}"
                            return str(x)

                        df_show = df_show.applymap(_fmt_cell)
                        st.dataframe(df_show, use_container_width=True)

    # ========================================
    # AI Summary Generation
    # ========================================
    try:
        # Get API key from Streamlit secrets or environment
        api_key = None
        try:
            api_key = st.secrets.get("ANTHROPIC_API_KEY")
        except:
            api_key = os.environ.get("ANTHROPIC_API_KEY")
        
        # Determine data availability
        has_balance_sheet = (tb_df is not None)
        has_cash_flow = (tb_df is not None and len(financial_data) >= 2)
        
        # Generate summary
        summary_text, used_ai = generate_ai_summary(
            financial_data=financial_data,
            has_balance_sheet=has_balance_sheet,
            has_cash_flow=has_cash_flow,
            api_key=api_key
        )
        
        # Display summary
        st.subheader("📊 AI Financial Summary")
        if used_ai:
            st.success("✅ Generated using Claude AI")
        else:
            st.info("ℹ️ Generated using rule-based analysis (AI unavailable)")
        
        st.markdown(summary_text)
        
    except Exception as e:
        st.warning(f"Could not generate AI summary: {str(e)}")
        summary_text = None
    
    # ========================================
    # PDF Report Generation
    # ========================================
    try:
        # Convert unit_scale to unit_label string
        unit_scale = st.session_state.get("unit_scale", 1000)
        if unit_scale == 1000:
            unit_label = "USD thousands"
        else:
            unit_label = "USD"
        
        pdf_bytes = generate_pdf_report(
            financial_data=financial_data,
            ai_summary=summary_text if summary_text else "Summary not available",
            unit_label=unit_label
        )
        
        st.download_button(
            "📄 Download PDF Report",
            data=pdf_bytes,
            file_name="financial_report.pdf",
            mime="application/pdf",
        )
        
    except Exception as e:
        st.warning(f"Could not generate PDF: {str(e)}")


# Persisted output preview / download (survives reruns)
if st.session_state.get("last_excel_bytes"):
    st.download_button(
        "Download Excel Output (last run)",
        data=st.session_state["last_excel_bytes"],
        file_name="3statement_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_last_excel",
    )